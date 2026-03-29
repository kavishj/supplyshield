"""
Run this script once to generate the demo supplier Excel:
    python generate_supplier_excel.py
Output: data/NexaCore_Suppliers_Demo.xlsx
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

OUT = Path("data/NexaCore_Suppliers_Demo.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

# ── Column definitions (header → field key) ──────────────────────────────────
COLUMNS = [
    "Supplier Name",
    "Country",
    "What They Supply",
    "Criticality Level",
    "Supply Category",
    "Annual Spend USD",
    "Spend % of Total Budget",
    "Contract Expiry Date",
    "Tier Level",
    "Sole Source",
    "On Time Delivery Rate",
    "Years In Relationship",
    "Financial Health",
    "Notes",
    "Order Fill Rate",
    "Audit Pass Rate",
    "Improvement Index",
    "Disruption Frequency",
    "Lead Time Variability",
    "Cyber Posture",
    "Inventory Buffer Days",
    "Has RTO Defined",
]

# ── 14 Suppliers ──────────────────────────────────────────────────────────────
# Expected risk outcomes:
#  1  Texas Instruments           → LOW  → AUTO_APPROVED
#  2  Murata Electronics NA       → LOW  → AUTO_APPROVED
#  3  TE Connectivity             → LOW  → AUTO_APPROVED
#  4  Vishay Dale Electronics     → LOW  → AUTO_APPROVED
#  5  Molex LLC                   → LOW  → AUTO_APPROVED
#  6  TDK Corporation of America  → LOW  → AUTO_APPROVED
#  7  Shenzhen Kinwong Electronic → MEDIUM → AUTO_APPROVED  (geo + delivery)
#  8  Foxconn Industrial Internet → MEDIUM → AUTO_APPROVED  (geo + news + delivery)
#  9  Advanced Semiconductor Eng. → MEDIUM → AUTO_APPROVED  (geo Taiwan Strait)
# 10  MediaTek Inc.               → MEDIUM → AUTO_APPROVED  (geo Taiwan)
# 11  Huawei Technologies         → HIGH  → REQUIRES_APPROVAL (Entity List, sole source)
# 12  ZTE Corporation             → HIGH  → REQUIRES_APPROVAL (sanctions history, news)
# 13  Iran Electronics Industries → BLOCKED (OFAC SDN direct match)
# 14  Renova Industrial Solutions → BLOCKED (OFAC 50% Rule — Viktor Vekselberg ~70%)

SUPPLIERS = [
    # ── 1 LOW ────────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "Texas Instruments Inc.",
        "Country":                 "UNITED STATES",
        "What They Supply":        "Analog ICs, microcontrollers, DSPs, and power management semiconductors for industrial control systems",
        "Criticality Level":       "Critical",
        "Supply Category":         "Semiconductors & ICs",
        "Annual Spend USD":        4200000,
        "Spend % of Total Budget": 12.8,
        "Contract Expiry Date":    "2028-12-31",
        "Tier Level":              "Tier 1",
        "Sole Source":             0,
        "On Time Delivery Rate":   97.5,
        "Years In Relationship":   12,
        "Financial Health":        "Good",
        "Notes":                   "Strategic Tier 1 semiconductor partner. Dual-qualified with Analog Devices for critical power ICs. Dedicated FAE support. ISO 26262 and AEC-Q100 certified product lines. Preferred supplier for all analog and mixed-signal ICs.",
        "Order Fill Rate":         97.0,
        "Audit Pass Rate":         99.0,
        "Improvement Index":       95.0,
        "Disruption Frequency":    0,
        "Lead Time Variability":   "Low",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   60,
        "Has RTO Defined":         1,
    },
    # ── 2 LOW ────────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "Murata Electronics North America Inc.",
        "Country":                 "UNITED STATES",
        "What They Supply":        "MLCC capacitors, EMI filters, inductors, and RF components (0402–1210 series) for power conversion modules",
        "Criticality Level":       "Critical",
        "Supply Category":         "Electronic Components",
        "Annual Spend USD":        3100000,
        "Spend % of Total Budget": 9.5,
        "Contract Expiry Date":    "2027-06-30",
        "Tier Level":              "Tier 1",
        "Sole Source":             0,
        "On Time Delivery Rate":   98.2,
        "Years In Relationship":   9,
        "Financial Health":        "Good",
        "Notes":                   "Primary MLCC and passive component partner. Dual-qualified alongside TDK for all capacitor families ≥0402. Dedicated account manager and VMI programme in place. Zero compliance concerns in 9-year relationship.",
        "Order Fill Rate":         98.0,
        "Audit Pass Rate":         99.0,
        "Improvement Index":       97.0,
        "Disruption Frequency":    0,
        "Lead Time Variability":   "Low",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   45,
        "Has RTO Defined":         1,
    },
    # ── 3 LOW ────────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "TE Connectivity Ltd.",
        "Country":                 "UNITED STATES",
        "What They Supply":        "Board-to-board connectors, industrial-grade wire harnesses, and sealed connector assemblies for harsh-environment ECUs",
        "Criticality Level":       "High",
        "Supply Category":         "Connectors & Interconnects",
        "Annual Spend USD":        2400000,
        "Spend % of Total Budget": 7.3,
        "Contract Expiry Date":    "2027-09-30",
        "Tier Level":              "Tier 1",
        "Sole Source":             0,
        "On Time Delivery Rate":   96.0,
        "Years In Relationship":   7,
        "Financial Health":        "Good",
        "Notes":                   "Primary connector supplier for all defense-grade ECU programs. Products qualified to MIL-DTL-38999 Series III. Secondary source: Amphenol for board-to-board connectors. No quality escapes in 7 years.",
        "Order Fill Rate":         96.0,
        "Audit Pass Rate":         97.0,
        "Improvement Index":       94.0,
        "Disruption Frequency":    0,
        "Lead Time Variability":   "Low",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   45,
        "Has RTO Defined":         1,
    },
    # ── 4 LOW ────────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "Vishay Dale Electronics LLC",
        "Country":                 "UNITED STATES",
        "What They Supply":        "Precision resistors, film capacitors, and power inductors for high-reliability PCB assemblies",
        "Criticality Level":       "Medium",
        "Supply Category":         "Electronic Components",
        "Annual Spend USD":        980000,
        "Spend % of Total Budget": 3.0,
        "Contract Expiry Date":    "2027-03-31",
        "Tier Level":              "Tier 2",
        "Sole Source":             0,
        "On Time Delivery Rate":   95.0,
        "Years In Relationship":   6,
        "Financial Health":        "Good",
        "Notes":                   "Approved vendor for MIL-PRF-55342 and DSCC-qualified resistor families. Used for all Class 3 assemblies requiring IPC-A-610 compliance. Backup: Ohmite Mfg. for high-power resistors.",
        "Order Fill Rate":         95.0,
        "Audit Pass Rate":         96.0,
        "Improvement Index":       93.0,
        "Disruption Frequency":    0,
        "Lead Time Variability":   "Low",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   30,
        "Has RTO Defined":         1,
    },
    # ── 5 LOW ────────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "Molex LLC",
        "Country":                 "UNITED STATES",
        "What They Supply":        "High-speed data connectors, FFC/FPC assemblies, and power connectors for industrial automation control boards",
        "Criticality Level":       "Medium",
        "Supply Category":         "Connectors & Interconnects",
        "Annual Spend USD":        750000,
        "Spend % of Total Budget": 2.3,
        "Contract Expiry Date":    "2026-12-31",
        "Tier Level":              "Tier 2",
        "Sole Source":             0,
        "On Time Delivery Rate":   94.0,
        "Years In Relationship":   5,
        "Financial Health":        "Good",
        "Notes":                   "Preferred supplier for Mini-Fit Jr. and Micro-Fit connector families. Contract renewal due December 2026 — initiate negotiations Q3 2026. One minor delivery delay in Q1 2026 (resolved). Cyber audit completed 2025.",
        "Order Fill Rate":         94.0,
        "Audit Pass Rate":         94.0,
        "Improvement Index":       92.0,
        "Disruption Frequency":    1,
        "Lead Time Variability":   "Low",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   30,
        "Has RTO Defined":         1,
    },
    # ── 6 LOW ────────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "TDK Corporation of America",
        "Country":                 "UNITED STATES",
        "What They Supply":        "Ferrite cores, common-mode chokes, power inductors, and EMC components for power conversion modules",
        "Criticality Level":       "High",
        "Supply Category":         "Electronic Components",
        "Annual Spend USD":        1800000,
        "Spend % of Total Budget": 5.5,
        "Contract Expiry Date":    "2028-03-31",
        "Tier Level":              "Tier 1",
        "Sole Source":             0,
        "On Time Delivery Rate":   96.5,
        "Years In Relationship":   8,
        "Financial Health":        "Good",
        "Notes":                   "Dual-qualified alongside Murata for MLCC and choke families. Strong JIT programme with 6-week rolling forecast. TDK parent (Japan) provides supply continuity backing. No compliance or quality issues.",
        "Order Fill Rate":         96.0,
        "Audit Pass Rate":         98.0,
        "Improvement Index":       95.0,
        "Disruption Frequency":    0,
        "Lead Time Variability":   "Low",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   45,
        "Has RTO Defined":         1,
    },
    # ── 7 MEDIUM ─────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "Shenzhen Kinwong Electronic Co. Ltd",
        "Country":                 "CHINA",
        "What They Supply":        "6-layer and 8-layer HDI PCBs for industrial motor drive controllers and power inverter boards",
        "Criticality Level":       "High",
        "Supply Category":         "PCB Fabrication",
        "Annual Spend USD":        1600000,
        "Spend % of Total Budget": 4.9,
        "Contract Expiry Date":    "2026-08-31",
        "Tier Level":              "Tier 2",
        "Sole Source":             0,
        "On Time Delivery Rate":   82.0,
        "Years In Relationship":   3,
        "Financial Health":        "Fair",
        "Notes":                   "Cost-competitive HDI PCB vendor for industrial motor drive programs. OTD slipped from 91% (2024) to 82% (Q1 2026) due to raw material shortages in Shenzhen. Geographic concentration risk. Contract renewal negotiations ongoing. Seek Malaysia backup qualification — TTM Technologies shortlisted.",
        "Order Fill Rate":         80.0,
        "Audit Pass Rate":         78.0,
        "Improvement Index":       65.0,
        "Disruption Frequency":    2,
        "Lead Time Variability":   "Medium",
        "Cyber Posture":           "Fair",
        "Inventory Buffer Days":   20,
        "Has RTO Defined":         0,
    },
    # ── 8 MEDIUM ─────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "Foxconn Industrial Internet Co. Ltd",
        "Country":                 "CHINA",
        "What They Supply":        "Contract electronics manufacturing (EMS) — SMT assembly, box-build, and functional test for industrial IoT gateway units",
        "Criticality Level":       "Critical",
        "Supply Category":         "Electronics Manufacturing Services",
        "Annual Spend USD":        5800000,
        "Spend % of Total Budget": 17.7,
        "Contract Expiry Date":    "2026-11-30",
        "Tier Level":              "Tier 1",
        "Sole Source":             0,
        "On Time Delivery Rate":   79.0,
        "Years In Relationship":   4,
        "Financial Health":        "Fair",
        "Notes":                   "Primary EMS partner for IoT gateway manufacturing. OTD degraded significantly in H2 2025 due to COVID-related facility disruptions in Shenzhen. Adverse labour practice media coverage noted in Q4 2025. Dual EMS qualification with Jabil Circuit underway — targeting Q3 2026 completion.",
        "Order Fill Rate":         77.0,
        "Audit Pass Rate":         75.0,
        "Improvement Index":       60.0,
        "Disruption Frequency":    3,
        "Lead Time Variability":   "Medium",
        "Cyber Posture":           "Fair",
        "Inventory Buffer Days":   15,
        "Has RTO Defined":         0,
    },
    # ── 9 MEDIUM ─────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "Advanced Semiconductor Engineering Inc.",
        "Country":                 "TAIWAN",
        "What They Supply":        "Flip-chip and wire-bond IC packaging, substrate design, and final-test services for custom ASIC chipsets",
        "Criticality Level":       "Critical",
        "Supply Category":         "Semiconductors & ICs",
        "Annual Spend USD":        3400000,
        "Spend % of Total Budget": 10.4,
        "Contract Expiry Date":    "2027-04-30",
        "Tier Level":              "Tier 1",
        "Sole Source":             0,
        "On Time Delivery Rate":   88.0,
        "Years In Relationship":   5,
        "Financial Health":        "Good",
        "Notes":                   "Primary IC packaging partner for custom ASICs. Taiwan Strait geopolitical risk is primary concern — Business Continuity Plan requires dual-site qualification. Secondary packaging site at ASE Penang (Malaysia) partially qualified for wire-bond products only. Inventory buffer policy increased to 45 days in 2025.",
        "Order Fill Rate":         89.0,
        "Audit Pass Rate":         91.0,
        "Improvement Index":       85.0,
        "Disruption Frequency":    1,
        "Lead Time Variability":   "Medium",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   45,
        "Has RTO Defined":         1,
    },
    # ── 10 MEDIUM ────────────────────────────────────────────────────────────
    {
        "Supplier Name":           "MediaTek Inc.",
        "Country":                 "TAIWAN",
        "What They Supply":        "Baseband SoCs, Wi-Fi 6E chipsets, and Bluetooth combo ICs for industrial IoT edge compute modules",
        "Criticality Level":       "High",
        "Supply Category":         "Semiconductors & ICs",
        "Annual Spend USD":        2100000,
        "Spend % of Total Budget": 6.4,
        "Contract Expiry Date":    "2027-07-31",
        "Tier Level":              "Tier 2",
        "Sole Source":             0,
        "On Time Delivery Rate":   91.0,
        "Years In Relationship":   3,
        "Financial Health":        "Good",
        "Notes":                   "Preferred wireless SoC vendor for next-gen IoT edge modules. Strong roadmap alignment with NexaCore's 5G-IoT programme. Taiwan Strait risk acknowledged — Qualcomm evaluated as alternative source for Wi-Fi 6E ICs. Lead time increased from 16 to 20 weeks in 2025 due to TSMC capacity constraints.",
        "Order Fill Rate":         88.0,
        "Audit Pass Rate":         89.0,
        "Improvement Index":       82.0,
        "Disruption Frequency":    1,
        "Lead Time Variability":   "Medium",
        "Cyber Posture":           "Good",
        "Inventory Buffer Days":   35,
        "Has RTO Defined":         1,
    },
    # ── 11 HIGH → REQUIRES_APPROVAL ──────────────────────────────────────────
    # Run via SINGLE analysis with: Financial Health=Poor, OTD=65%, Lead Time=28wk,
    # Sole Source=Yes, Contract Expiry=2026-04-16 → score ≈ 0.76 (HIGH)
    {
        "Supplier Name":           "Huawei Technologies Co. Ltd",
        "Country":                 "CHINA",
        "What They Supply":        "Custom ASIC chipsets and 5G baseband modules for industrial IoT gateway units (part #NC-5G-GW-1024)",
        "Criticality Level":       "Critical",
        "Supply Category":         "Semiconductors & ICs",
        "Annual Spend USD":        1850000,
        "Spend % of Total Budget": 5.7,
        "Contract Expiry Date":    "2026-04-16",
        "Tier Level":              "Tier 1",
        "Sole Source":             1,
        "On Time Delivery Rate":   65.0,
        "Years In Relationship":   4,
        "Financial Health":        "Poor",
        "Notes":                   "SOLE-QUALIFIED supplier for NexaCore's IoT gateway ASIC (part #NC-5G-GW-1024). US Commerce Dept Entity List restrictions have extended lead times to 28 weeks. Contract lapsing — renewal stalled pending legal review of export licence implications under EAR Part 744. OTD degraded from 91% (2024) to 65% (2025) due to chip allocation delays under Huawei's HiSilicon supply constraints. Payment terms renegotiated unfavourably. HIGH risk — requires procurement manager sign-off.",
        "Order Fill Rate":         62.0,
        "Audit Pass Rate":         60.0,
        "Improvement Index":       45.0,
        "Disruption Frequency":    4,
        "Lead Time Variability":   "High",
        "Cyber Posture":           "Fair",
        "Inventory Buffer Days":   10,
        "Has RTO Defined":         0,
    },
    # ── 12 HIGH → REQUIRES_APPROVAL ──────────────────────────────────────────
    {
        "Supplier Name":           "ZTE Corporation",
        "Country":                 "CHINA",
        "What They Supply":        "5G NR radio frequency front-end modules and telecom-grade filters for industrial private network infrastructure",
        "Criticality Level":       "High",
        "Supply Category":         "RF & Electronic Systems",
        "Annual Spend USD":        1200000,
        "Spend % of Total Budget": 3.7,
        "Contract Expiry Date":    "2026-06-30",
        "Tier Level":              "Tier 2",
        "Sole Source":             0,
        "On Time Delivery Rate":   70.0,
        "Years In Relationship":   2,
        "Financial Health":        "Fair",
        "Notes":                   "RF front-end supplier for private 5G NR programmes. ZTE was subject to US Bureau of Industry and Security export restrictions (2018–2019, lifted). Ongoing adverse media coverage related to national security concerns and foreign government influence. OTD degraded to 70%. Cyber posture assessment rated Poor following third-party pen test in Q4 2025. Recommend dual-source qualification with Qorvo Inc. for RF filter modules.",
        "Order Fill Rate":         68.0,
        "Audit Pass Rate":         65.0,
        "Improvement Index":       50.0,
        "Disruption Frequency":    3,
        "Lead Time Variability":   "High",
        "Cyber Posture":           "Poor",
        "Inventory Buffer Days":   10,
        "Has RTO Defined":         0,
    },
    # ── 13 BLOCKED — OFAC SDN DIRECT MATCH ───────────────────────────────────
    {
        "Supplier Name":           "Iran Electronics Industries",
        "Country":                 "IRAN",
        "What They Supply":        "RF subsystem modules and military-grade connectors for defense-adjacent ECU programs",
        "Criticality Level":       "High",
        "Supply Category":         "RF & Electronic Systems",
        "Annual Spend USD":        480000,
        "Spend % of Total Budget": 1.5,
        "Contract Expiry Date":    "2026-09-30",
        "Tier Level":              "Tier 2",
        "Sole Source":             0,
        "On Time Delivery Rate":   78.0,
        "Years In Relationship":   1,
        "Financial Health":        "Fair",
        "Notes":                   "FLAGGED FOR COMPLIANCE REVIEW. Introduced through Singapore-based intermediary (SingTech Distribution Pte Ltd) as a non-disclosed principal. SupplyShield OFAC screening identified this as Iran Electronics Industries — a designated entity on the OFAC SDN list under the Iran, IRGC, and WMD programmes. Procurement engagement HALTED. Legal team notified. All POs cancelled. Do not re-engage without General Counsel approval.",
        "Order Fill Rate":         72.0,
        "Audit Pass Rate":         55.0,
        "Improvement Index":       40.0,
        "Disruption Frequency":    2,
        "Lead Time Variability":   "High",
        "Cyber Posture":           "Poor",
        "Inventory Buffer Days":   0,
        "Has RTO Defined":         0,
    },
    # ── 14 BLOCKED — OFAC 50% RULE ───────────────────────────────────────────
    {
        "Supplier Name":           "Renova Industrial Solutions",
        "Country":                 "RUSSIA",
        "What They Supply":        "Titanium alloy heat sinks and precision-machined aluminium enclosures for high-power electronics",
        "Criticality Level":       "Medium",
        "Supply Category":         "Mechanical Components",
        "Annual Spend USD":        720000,
        "Spend % of Total Budget": 2.2,
        "Contract Expiry Date":    "2027-03-31",
        "Tier Level":              "Tier 3",
        "Sole Source":             0,
        "On Time Delivery Rate":   84.0,
        "Years In Relationship":   1,
        "Financial Health":        "Fair",
        "Notes":                   "Russian subsidiary of Renova Group. Competitive titanium machining capability introduced via Hannover Messe 2025 referral. Beneficial ownership verification PENDING at time of onboarding. SupplyShield OFAC 50% Rule check identified Viktor Vekselberg (OFAC SDN — EO 13661, Russia/Ukraine programme) as majority beneficial owner (~70%). Cumulative OFAC stake ≥ 50% threshold — entity is BLOCKED under OFAC 50% Rule. Procurement engagement HALTED. Legal escalation required.",
        "Order Fill Rate":         75.0,
        "Audit Pass Rate":         62.0,
        "Improvement Index":       48.0,
        "Disruption Frequency":    2,
        "Lead Time Variability":   "Medium",
        "Cyber Posture":           "Poor",
        "Inventory Buffer Days":   0,
        "Has RTO Defined":         0,
    },
]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Suppliers"

# Colours
C_HEADER_BG  = "1A1F35"
C_HEADER_FG  = "E2E8F0"
C_ACCENT     = "6C63FF"
C_LOW        = "D1FAE5"
C_MEDIUM     = "FEF9C3"
C_HIGH       = "FEE2E2"
C_BLOCKED    = "F3E8FF"
C_ALT        = "F8F9FB"
C_WHITE      = "FFFFFF"
C_BORDER     = "CBD5E1"

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
title_cell = ws["A1"]
title_cell.value = "NexaCore Systems Inc. — Supplier Onboarding Register"
title_cell.font      = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
title_cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── Sub-title row ─────────────────────────────────────────────────────────────
ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
sub = ws["A2"]
sub.value      = "Defense Electronics & Industrial Automation  ·  Import-ready for SupplyShield  ·  14 suppliers covering all risk tiers"
sub.font       = Font(name="Calibri", italic=True, size=10, color="94A3B8")
sub.fill       = PatternFill("solid", fgColor=C_HEADER_BG)
sub.alignment  = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Header row ────────────────────────────────────────────────────────────────
for col_idx, col_name in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=3, column=col_idx, value=col_name)
    cell.font      = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    cell.fill      = PatternFill("solid", fgColor=C_ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER
ws.row_dimensions[3].height = 32

# ── Data rows ─────────────────────────────────────────────────────────────────
ROW_FILLS = [
    "D1FAE5",  # 1 LOW  (green tint)
    "D1FAE5",  # 2 LOW
    "D1FAE5",  # 3 LOW
    "D1FAE5",  # 4 LOW
    "D1FAE5",  # 5 LOW
    "D1FAE5",  # 6 LOW
    "FEF9C3",  # 7 MEDIUM (yellow tint)
    "FEF9C3",  # 8 MEDIUM
    "FEF9C3",  # 9 MEDIUM
    "FEF9C3",  # 10 MEDIUM
    "FEE2E2",  # 11 HIGH (red tint)
    "FEE2E2",  # 12 HIGH
    "F3E8FF",  # 13 BLOCKED-SDN (purple tint)
    "F3E8FF",  # 14 BLOCKED-50% Rule
]

for row_idx, supplier in enumerate(SUPPLIERS, start=4):
    fill_color = ROW_FILLS[row_idx - 4]
    fill       = PatternFill("solid", fgColor=fill_color)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        val  = supplier.get(col_name, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill      = fill
        cell.border    = BORDER
        cell.font      = Font(name="Calibri", size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 3 or col_idx == 14))

    ws.row_dimensions[row_idx].height = 40

# ── Column widths ─────────────────────────────────────────────────────────────
WIDTHS = [
    38,   # Supplier Name
    18,   # Country
    52,   # What They Supply
    16,   # Criticality Level
    28,   # Supply Category
    18,   # Annual Spend USD
    22,   # Spend %
    20,   # Contract Expiry Date
    12,   # Tier Level
    12,   # Sole Source
    22,   # On Time Delivery Rate
    22,   # Years In Relationship
    16,   # Financial Health
    55,   # Notes
    18,   # Order Fill Rate
    16,   # Audit Pass Rate
    18,   # Improvement Index
    22,   # Disruption Frequency
    22,   # Lead Time Variability
    14,   # Cyber Posture
    22,   # Inventory Buffer Days
    16,   # Has RTO Defined
]
for i, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Legend sheet ──────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend")
ls.column_dimensions["A"].width = 28
ls.column_dimensions["B"].width = 55

legend_rows = [
    ("FIELD", "ACCEPTED VALUES / FORMAT", C_HEADER_BG, C_HEADER_FG),
    ("Criticality Level",       "Critical / High / Medium / Low",                        "E2E8F0", "1E293B"),
    ("Supply Category",         "Semiconductors & ICs / PCB Fabrication / Electronic Components / Connectors & Interconnects / RF & Electronic Systems / Electronics Manufacturing Services / Mechanical Components / Other", "E2E8F0", "1E293B"),
    ("Tier Level",              "Tier 1 / Tier 2 / Tier 3",                              "E2E8F0", "1E293B"),
    ("Sole Source",             "1 = Yes, 0 = No",                                       "E2E8F0", "1E293B"),
    ("On Time Delivery Rate",   "0–100 (percentage, e.g. 95.5)",                         "E2E8F0", "1E293B"),
    ("Financial Health",        "Good / Fair / Poor",                                    "E2E8F0", "1E293B"),
    ("Order Fill Rate",         "0–100 (OTIF in-full %, e.g. 92)",                       "E2E8F0", "1E293B"),
    ("Audit Pass Rate",         "0–100 (compliance audit pass %, e.g. 85)",              "E2E8F0", "1E293B"),
    ("Improvement Index",       "0–100 (corrective action closure %, e.g. 78)",          "E2E8F0", "1E293B"),
    ("Disruption Frequency",    "Integer — incidents per year (e.g. 2)",                 "E2E8F0", "1E293B"),
    ("Lead Time Variability",   "Low / Medium / High",                                   "E2E8F0", "1E293B"),
    ("Cyber Posture",           "Good / Fair / Poor",                                    "E2E8F0", "1E293B"),
    ("Inventory Buffer Days",   "Integer — days of supply on hand (e.g. 45)",            "E2E8F0", "1E293B"),
    ("Has RTO Defined",         "1 = Yes (RTO documented), 0 = No",                      "E2E8F0", "1E293B"),
    ("Contract Expiry Date",    "YYYY-MM-DD format (e.g. 2027-06-30)",                   "E2E8F0", "1E293B"),
    ("Annual Spend USD",        "Numeric USD value — no $ or commas (e.g. 1500000)",     "E2E8F0", "1E293B"),
    ("Spend % of Total Budget", "0–100 (e.g. 12.5)",                                     "E2E8F0", "1E293B"),
    ("", "", "FFFFFF", "000000"),
    ("RISK TIER COLOUR CODE",   "",                                                       C_HEADER_BG, C_HEADER_FG),
    ("GREEN rows",              "LOW risk → AUTO_APPROVED",                              "D1FAE5", "065F46"),
    ("YELLOW rows",             "MEDIUM risk → AUTO_APPROVED (risk flagged)",            "FEF9C3", "854D0E"),
    ("RED rows",                "HIGH risk → REQUIRES_APPROVAL",                         "FEE2E2", "991B1B"),
    ("PURPLE rows",             "BLOCKED — OFAC SDN direct match OR OFAC 50% Rule",      "F3E8FF", "6B21A8"),
]

for r, (field, desc, bg, fg) in enumerate(legend_rows, start=1):
    c1 = ls.cell(row=r, column=1, value=field)
    c2 = ls.cell(row=r, column=2, value=desc)
    for c in (c1, c2):
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(name="Calibri", bold=(r == 1 or field in ("RISK TIER COLOUR CODE",)), size=9, color=fg)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border    = BORDER
    ls.row_dimensions[r].height = 22

wb.save(OUT)
print(f"Excel saved -> {OUT}")
print(f"  {len(SUPPLIERS)} suppliers across all risk tiers")
print(f"  Columns: {len(COLUMNS)}")
